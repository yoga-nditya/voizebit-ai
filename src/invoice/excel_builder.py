import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from config_new import FILES_DIR


def _side(style="thin"):
    return Side(style=style, color="000000")


def apply_outer_and_vertical_only(ws, r1, c1, r2, c2, vertical_separators_cols, outer_style="medium", inner_style="thin"):
    outer = _side(outer_style)
    inner = _side(inner_style)
    seps = set(vertical_separators_cols or [])

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left = None
            right = None
            top = None
            bottom = None

            if c == c1:
                left = outer
            if c == c2:
                right = outer
            if r == r1:
                top = outer
            if r == r2:
                bottom = outer

            if c in seps and c != c1:
                left = inner if left is None else left
            if (c + 1) in seps and c != c2:
                right = inner if right is None else right

            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for r in range(r1, r2 + 1):
        cell = ws.cell(r, c2)
        cell.border = Border(
            left=cell.border.left,
            right=outer,
            top=cell.border.top,
            bottom=cell.border.bottom
        )


def set_outer_border_only(ws, r1, c1, r2, c2, style="medium"):
    outer = _side(style)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left = outer if c == c1 else None
            right = outer if c == c2 else None
            top = outer if r == r1 else None
            bottom = outer if r == r2 else None
            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


def create_invoice_xlsx(inv: dict, fname_base: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 110

    bold = Font(bold=True)
    bold_ul = Font(bold=True, underline="single")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    right_mid = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def money(cell):
        cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 3

    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 19
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18

    payment = inv.get("payment") or {}
    defaults = {
        "beneficiary": "PT. Sarana Trans Bersama Jaya",
        "bank_name": "BCA",
        "branch": "Cibadak - Sukabumi",
        "idr_acct": "35212 26666",
    }
    for k, v in defaults.items():
        if not payment.get(k):
            payment[k] = v

    invoice_no = inv.get("invoice_no") or ""
    inv_date = inv.get("invoice_date") or datetime.now().strftime("%d-%b-%y")

    bill_to = inv.get("bill_to") or {}
    ship_to = inv.get("ship_to") or {}
    attn = inv.get("attn") or "Accounting / Finance"
    phone = inv.get("phone") or ""
    fax = inv.get("fax") or ""

    sales_person = inv.get("sales_person") or "Syaeful Bakri"
    ref_no = inv.get("ref_no") or ""
    ship_via = inv.get("ship_via") or ""
    ship_date = inv.get("ship_date") or ""
    terms = inv.get("terms") or ""
    no_surat_jalan = inv.get("no_surat_jalan") or ""

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 18
    ws.row_dimensions[14].height = 18

    ws["D1"].value = "Bill To:"
    ws["D1"].font = bold
    ws.merge_cells("D1:F1")

    ws["H1"].value = "Ship To:"
    ws["H1"].font = bold
    ws.merge_cells("H1:J1")

    bill_text = "\n".join([x for x in [
        (bill_to.get("name") or "").strip(),
        (bill_to.get("address") or "").strip(),
        (bill_to.get("address2") or "").strip()
    ] if x])

    ship_text = "\n".join([x for x in [
        (ship_to.get("name") or "").strip(),
        (ship_to.get("address") or "").strip(),
        (ship_to.get("address2") or "").strip()
    ] if x])

    ws["D2"].value = bill_text
    ws.merge_cells("D2:F3")
    ws["D2"].alignment = left

    ws["H2"].value = ship_text
    ws.merge_cells("H2:J3")
    ws["H2"].alignment = left

    ws["D5"].value = "Phone:"
    ws["D5"].font = bold
    ws.merge_cells("E5:F5")
    ws["E5"].value = phone
    ws["E5"].alignment = left_mid

    ws["H5"].value = "Fax:"
    ws["H5"].font = bold
    ws.merge_cells("I5:J5")
    ws["I5"].value = fax
    ws["I5"].alignment = left_mid

    ws["D7"].value = "Attn :"
    ws["D7"].font = bold
    ws.merge_cells("E7:F7")
    ws["E7"].value = attn
    ws["E7"].alignment = left_mid

    ws["I6"].value = "Invoice"
    ws["I6"].font = bold
    ws["I6"].alignment = right_mid
    ws["J6"].value = invoice_no
    ws["J6"].alignment = left_mid

    ws["I7"].value = "Date"
    ws["I7"].font = bold
    ws["I7"].alignment = right_mid
    ws["J7"].value = inv_date
    ws["J7"].alignment = left_mid

    ws["I8"].value = "No. Surat Jalan"
    ws["I8"].font = bold
    ws["I8"].alignment = right_mid
    ws["J8"].value = no_surat_jalan
    ws["J8"].alignment = left_mid

    ws.merge_cells("D10:E10")
    ws["D10"].value = "Ref No."
    ws["D10"].font = bold
    ws["D10"].alignment = center

    ws.merge_cells("F10:G10")
    ws["F10"].value = "Sales Person"
    ws["F10"].font = bold
    ws["F10"].alignment = center

    ws["H10"].value = "Ship Via"
    ws["H10"].font = bold
    ws["H10"].alignment = center

    ws["I10"].value = "Ship Date"
    ws["I10"].font = bold
    ws["I10"].alignment = center

    ws["J10"].value = "Terms"
    ws["J10"].font = bold
    ws["J10"].alignment = center

    ws.merge_cells("D11:E13")
    ws["D11"].value = ref_no
    ws["D11"].alignment = center

    ws.merge_cells("F11:G13")
    ws["F11"].value = sales_person
    ws["F11"].alignment = center

    ws.merge_cells("H11:H13")
    ws["H11"].value = ship_via
    ws["H11"].alignment = center

    ws.merge_cells("I11:I13")
    ws["I11"].value = ship_date
    ws["I11"].alignment = center

    ws.merge_cells("J11:J13")
    ws["J11"].value = terms
    ws["J11"].alignment = center

    apply_outer_and_vertical_only(ws, 10, 4, 13, 10, vertical_separators_cols=[6, 8, 9, 10])

    ws["D14"].value = "Qty"
    ws["E14"].value = ""
    ws["F14"].value = "Date"
    ws.merge_cells("G14:H14")
    ws["G14"].value = "Description"
    ws["I14"].value = "Price"
    ws["J14"].value = "Amount (IDR)"

    for c in ["D", "E", "F", "G", "I", "J"]:
        ws[f"{c}14"].font = bold
        ws[f"{c}14"].alignment = center
    ws["H14"].alignment = center

    items = inv.get("items") or []
    start_row = 15
    max_rows = max(10, len(items))
    subtotal = 0

    for idx in range(max_rows):
        r = start_row + idx
        ws.merge_cells(f"G{r}:H{r}")

        ws[f"D{r}"].alignment = center
        ws[f"E{r}"].alignment = center
        ws[f"F{r}"].alignment = center
        ws[f"G{r}"].alignment = left
        ws[f"I{r}"].alignment = right
        ws[f"J{r}"].alignment = right

        if idx < len(items):
            it = items[idx]
            qty = float(it.get("qty") or 0)
            unit = (it.get("unit") or "Kg").strip()
            dt = it.get("date") or inv_date
            desc = (it.get("description") or "").strip()
            price = int(it.get("price") or 0)
            amount = int(round(qty * price))
            subtotal += amount

            ws[f"D{r}"].value = qty if qty % 1 != 0 else int(qty)
            ws[f"E{r}"].value = unit
            ws[f"F{r}"].value = dt
            ws[f"G{r}"].value = desc
            ws[f"I{r}"].value = price
            ws[f"J{r}"].value = amount
            money(ws[f"I{r}"])
            money(ws[f"J{r}"])

    last_table_row = start_row + max_rows - 1
    apply_outer_and_vertical_only(ws, 14, 4, last_table_row, 10, vertical_separators_cols=[5, 6, 7, 9, 10])

    freight = int(inv.get("freight") or 0)
    ppn_rate = float(inv.get("ppn_rate") or 0.11)
    deposit = int(inv.get("deposit") or 0)

    total_before_ppn = subtotal + freight
    ppn = int(round(total_before_ppn * ppn_rate))
    balance = total_before_ppn + ppn - deposit

    base_row = last_table_row + 2

    ws.merge_cells(f"D{base_row}:H{base_row}")
    ws[f"D{base_row}"].value = "Please Transfer Full Amount to:"
    ws[f"D{base_row}"].font = bold_ul
    ws[f"D{base_row}"].alignment = left_mid

    left_lines = [
        f"Beneficiary : {payment.get('beneficiary','')}",
        f"Bank Name   : {payment.get('bank_name','')}",
        f"Branch      : {payment.get('branch','')}",
        f"IDR Acct    : {payment.get('idr_acct','')}",
    ]
    for i, line in enumerate(left_lines, start=1):
        ws.merge_cells(f"D{base_row+i}:H{base_row+i}")
        ws[f"D{base_row+i}"].value = line
        ws[f"D{base_row+i}"].alignment = left_mid

    labels = [
        ("Total", subtotal),
        ("Freight", freight),
        ("Total", total_before_ppn),
        (f"PPN {int(ppn_rate*100)}%", ppn),
        ("Less: Deposit", deposit),
        ("Balance Due", balance),
    ]
    totals_top = base_row

    thin_side = Side(style="thin", color="000000")
    no_border = Border()

    def border_box_thin():
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for i, (lab, val) in enumerate(labels):
        rr = totals_top + i

        ws[f"I{rr}"].value = lab
        ws[f"I{rr}"].alignment = right
        ws[f"I{rr}"].font = Font(bold=True)
        ws[f"I{rr}"].border = no_border

        ws[f"J{rr}"].value = val
        ws[f"J{rr}"].alignment = right
        ws[f"J{rr}"].font = Font(bold=True) if lab in ("Balance Due",) else Font(bold=False)
        ws[f"J{rr}"].border = border_box_thin()
        money(ws[f"J{rr}"])

    totals_bottom = totals_top + len(labels) - 1

    box_top = totals_bottom + 2
    box_bottom = box_top + 6

    ws.merge_cells(f"G{box_top}:J{box_top}")
    ws[f"G{box_top}"].value = "PT. Sarana Trans Bersama Jaya"
    ws[f"G{box_top}"].alignment = center
    ws[f"G{box_top}"].font = Font(bold=True)

    set_outer_border_only(ws, box_top, 7, box_bottom, 10, style="medium")

    footer_row = box_bottom + 1
    ws.merge_cells(f"G{footer_row}:J{footer_row}")
    ws[f"G{footer_row}"].value = "Please kindly fax to our attention upon receipt"
    ws[f"G{footer_row}"].alignment = center

    try:
        folder = str(FILES_DIR)
    except Exception:
        folder = "static/files"
    os.makedirs(folder, exist_ok=True)

    out_path = os.path.join(folder, f"{fname_base}.xlsx")
    wb.save(out_path)
    return f"{fname_base}.xlsx"
