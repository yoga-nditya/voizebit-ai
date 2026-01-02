import os
import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from limbah_database import (
    find_limbah_by_kode,
    find_limbah_by_jenis,
    convert_voice_to_number,
)
from utils import (
    db_insert_history, db_append_message, db_update_state,
    search_company_address, search_company_address_ai,
)
from config_new import FILES_DIR


def is_non_b3_input(text: str) -> bool:
    if not text:
        return False
    t = text.strip().lower()
    norm = re.sub(r"[\s\-_]+", "", t)
    return norm in ("nonb3", "nonbii3") or norm.startswith("nonb3")


def normalize_id_number_text(text: str) -> str:
    if not text:
        return text
    t = text.strip()
    t = re.sub(r"(?<=\d)[\.,](?=\d{3}(\D|$))", "", t)
    t = re.sub(r"(?<=\d),(?=\d)", ".", t)
    return t


# =========================
# Improved Indonesian words-to-number (supports koma, ribu, juta, dll)
# =========================
_ID_SMALL = {
    "nol": 0, "kosong": 0,
    "satu": 1, "se": 1,
    "dua": 2, "tiga": 3, "empat": 4, "lima": 5,
    "enam": 6, "tujuh": 7, "delapan": 8, "sembilan": 9,
    "sepuluh": 10, "sebelas": 11,
}

_ID_SCALES = {
    "ribu": 1_000,
    "juta": 1_000_000,
    "miliar": 1_000_000_000,
    "triliun": 1_000_000_000_000,
}


def _tokenize_id_words(s: str):
    s = (s or "").lower().strip()
    s = re.sub(r"[-_]", " ", s)
    s = re.sub(r"[^a-z0-9\s,\.]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.split()


def _parse_id_integer_words(tokens):
    total = 0
    current = 0
    i = 0
    while i < len(tokens):
        w = tokens[i]

        if re.fullmatch(r"\d+(?:\.\d+)?", w):
            try:
                current += int(float(w))
            except:
                pass
            i += 1
            continue

        if w == "seratus":
            current += 100
            i += 1
            continue

        if w == "seribu":
            total += (current if current else 1) * 1000
            current = 0
            i += 1
            continue

        if w in _ID_SMALL:
            val = _ID_SMALL[w]

            if i + 1 < len(tokens) and tokens[i + 1] == "belas":
                current += 10 + val
                i += 2
                continue

            if i + 1 < len(tokens) and tokens[i + 1] == "puluh":
                current += val * 10
                i += 2
                continue

            if i + 1 < len(tokens) and tokens[i + 1] == "ratus":
                current += val * 100
                i += 2
                continue

            current += val
            i += 1
            continue

        if w in _ID_SCALES:
            scale = _ID_SCALES[w]
            if current == 0:
                current = 1
            total += current * scale
            current = 0
            i += 1
            continue

        i += 1

    return total + current


def words_to_number_id(text: str):
    if not text:
        return None

    raw = text.strip().lower()
    norm = normalize_id_number_text(raw)
    if re.fullmatch(r"\d+(?:\.\d+)?", norm):
        try:
            return float(norm)
        except:
            pass

    tokens = _tokenize_id_words(raw)
    if not tokens:
        return None

    if "koma" in tokens:
        k = tokens.index("koma")
        left = tokens[:k]
        right = tokens[k + 1:]

        left_int = _parse_id_integer_words(left) if left else 0

        if not right:
            return float(left_int) + 0.5

        digits = []
        for w in right:
            if w in _ID_SMALL:
                digits.append(str(_ID_SMALL[w]))
                continue
            if re.fullmatch(r"\d+", w):
                digits.append(w)
                continue
            if w in _ID_SCALES:
                break

        frac_str = "".join(digits) if digits else "5"
        frac_val = float("0." + frac_str)

        scale = None
        for w in right:
            if w in _ID_SCALES:
                scale = _ID_SCALES[w]
                break

        val = float(left_int) + frac_val
        if scale:
            val *= scale
        return val

    return float(_parse_id_integer_words(tokens))


def parse_amount_id(text: str) -> int:
    if not text:
        return 0

    raw = text.strip().lower()

    wv = words_to_number_id(raw)
    if wv is not None:
        try:
            return int(round(float(wv)))
        except:
            pass

    tnorm = normalize_id_number_text(text)
    val = convert_voice_to_number(tnorm)
    if val is None:
        val = 0

    try:
        return int(round(float(val)))
    except:
        digits = re.sub(r"\D+", "", str(val))
        return int(digits) if digits else 0


def parse_qty_id(text: str) -> float:
    if not text:
        return 0.0

    raw = text.strip().lower()
    wv = words_to_number_id(raw)
    if wv is not None:
        try:
            return float(wv)
        except:
            pass

    t = normalize_id_number_text(text)
    v = convert_voice_to_number(t)
    try:
        return float(v)
    except:
        m = re.findall(r"\d+(?:\.\d+)?", t)
        return float(m[0]) if m else 0.0


def normalize_voice_strip(text: str) -> str:
    if not text:
        return text
    t = text.strip()
    low = t.lower()
    if low == "strip":
        return "-"
    if re.fullmatch(r"(.*\b)?strip(\b.*)?", low):
        return "-"
    return text


# =========================
# Address: kalau hasil search berisi kalimat panjang/penjelasan -> pakai "Di tempat"
# =========================
def _sanitize_company_address(addr: str) -> str:
    a = (addr or "").strip()
    if not a:
        return "Di tempat"

    low = a.lower()

    # ✅ FIX: deteksi jawaban AI panjang (seperti screenshot kamu) sebagai "not found"
    bad_patterns = [
        r"tidak\s*dapat\s+menentukan",
        r"tidak\s*bisa\s+menentukan",
        r"tidak\s*dapat\s+menemukan",
        r"tidak\s*bisa\s+menemukan",
        r"tidak\s*ditemukan",
        r"tidak\s*ketemu",
        r"tidak\s+ada\s+informasi",
        r"tidak\s+memiliki\s+informasi",
        r"saya\s+tidak\s+memiliki",
        r"informasi\s+yang\s+cukup",
        r"tidak\s+cukup\s+informasi",
        r"untuk\s+menentukan",
        r"nama\s+tersebut\s+terlalu\s+umum",
        r"terlalu\s+umum",
        r"tidak\s+spesifik",
        r"placeholder",
        r"nama\s+contoh",
        r"banyak\s+perusahaan.*nama\s+serupa",
        r"mungkin\s+menggunakan\s+nama\s+serupa",
        r"maaf",
        r"gagal",
        r"cannot\s+find",
        r"not\s+found",
        r"no\s+information",
        r"no\s+result",
        r"unknown",
    ]

    for p in bad_patterns:
        if re.search(p, low):
            return "Di tempat"

    # ✅ FIX: heuristik aman — kalau teks panjang kayak paragraf & tidak ada ciri alamat, jadikan "Di tempat"
    # (alamat biasanya ada: jl/jalan, rt/rw, kec/kab, kota, no., dsb)
    if len(a) > 120 and not re.search(r"\b(jl|jalan|rt|rw|kec|kel|kab|kota|no\.?|blok|desa)\b", low):
        return "Di tempat"

    # (opsional) kalau terlalu panjang banget tetap dianggap bukan alamat
    if len(a) > 250:
        return "Di tempat"

    return a


def resolve_company_address(company_name: str) -> str:
    addr = ""
    try:
        addr = (search_company_address(company_name) or "").strip()
    except:
        addr = ""
    addr = _sanitize_company_address(addr)
    if addr != "Di tempat":
        return addr

    try:
        addr2 = (search_company_address_ai(company_name) or "").strip()
    except:
        addr2 = ""
    return _sanitize_company_address(addr2)


def make_unique_filename_base(base_name: str) -> str:
    base_name = (base_name or "").strip()
    if not base_name:
        base_name = "Dokumen"
    try:
        folder = str(FILES_DIR)
    except Exception:
        folder = "static/files"

    def exists_any(name: str) -> bool:
        return (
            os.path.exists(os.path.join(folder, f"{name}.docx")) or
            os.path.exists(os.path.join(folder, f"{name}.pdf")) or
            os.path.exists(os.path.join(folder, f"{name}.xlsx")) or
            os.path.exists(os.path.join(folder, name))
        )

    if not exists_any(base_name):
        return base_name

    i = 2
    while True:
        candidate = f"{base_name} ({i})"
        if not exists_any(candidate):
            return candidate
        i += 1


def _invoice_counter_path() -> str:
    try:
        folder = str(FILES_DIR)
    except Exception:
        folder = "static/files"
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "invoice_counter.json")


def load_invoice_counter(prefix: str) -> int:
    path = _invoice_counter_path()
    try:
        if not os.path.exists(path):
            return 0
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        counters = data.get("counters") or {}
        return int(counters.get(prefix, 0))
    except:
        return 0


def save_invoice_counter(prefix: str, n: int) -> None:
    path = _invoice_counter_path()
    data = {}
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
    except:
        data = {}

    counters = data.get("counters") or {}
    counters[prefix] = int(n)
    data["counters"] = counters

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def get_next_invoice_no() -> str:
    now = datetime.now()
    prefix = now.strftime("%d%m")
    n = load_invoice_counter(prefix) + 1
    save_invoice_counter(prefix, n)
    return f"{prefix}{str(n).zfill(3)}"


def _side(style="thin"):
    return Side(style=style, color="000000")


def apply_outer_and_vertical_only(ws, r1, c1, r2, c2, vertical_separators_cols, outer_style="medium", inner_style="thin"):
    """
    Border:
    - outer = medium
    - vertical separators = thin
    - no horizontal inner borders
    """
    outer = _side(outer_style)
    inner = _side(inner_style)
    seps = set(vertical_separators_cols or [])

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left = None
            right = None
            top = None
            bottom = None

            if c == c1:
                left = outer
            if c == c2:
                right = outer
            if r == r1:
                top = outer
            if r == r2:
                bottom = outer

            if c in seps and c != c1:
                left = inner if left is None else left
            if (c + 1) in seps and c != c2:
                right = inner if right is None else right

            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)

    for r in range(r1, r2 + 1):
        cell = ws.cell(r, c2)
        cell.border = Border(
            left=cell.border.left,
            right=outer,
            top=cell.border.top,
            bottom=cell.border.bottom
        )


def set_outer_border_only(ws, r1, c1, r2, c2, style="medium"):
    outer = _side(style)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            left = outer if c == c1 else None
            right = outer if c == c2 else None
            top = outer if r == r1 else None
            bottom = outer if r == r2 else None
            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


# ==========================================
# Excel layout seperti template (gambar kedua)
# ==========================================
def create_invoice_xlsx(inv: dict, fname_base: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    ws.sheet_view.showGridLines = True
    ws.sheet_view.zoomScale = 110

    bold = Font(bold=True)
    bold_ul = Font(bold=True, underline="single")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    right_mid = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def money(cell):
        cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 3

    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 19
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 18

    payment = inv.get("payment") or {}
    defaults = {
        "beneficiary": "PT. Sarana Trans Bersama Jaya",
        "bank_name": "BCA",
        "branch": "Cibadak - Sukabumi",
        "idr_acct": "35212 26666",
    }
    for k, v in defaults.items():
        if not payment.get(k):
            payment[k] = v

    invoice_no = inv.get("invoice_no") or get_next_invoice_no()
    inv_date = inv.get("invoice_date") or datetime.now().strftime("%d-%b-%y")

    bill_to = inv.get("bill_to") or {}
    ship_to = inv.get("ship_to") or {}
    attn = inv.get("attn") or "Accounting / Finance"
    phone = inv.get("phone") or ""
    fax = inv.get("fax") or ""

    sales_person = inv.get("sales_person") or "Syaeful Bakri"
    ref_no = inv.get("ref_no") or ""
    ship_via = inv.get("ship_via") or ""
    ship_date = inv.get("ship_date") or ""
    terms = inv.get("terms") or ""
    no_surat_jalan = inv.get("no_surat_jalan") or ""

    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 34
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[10].height = 18
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 18
    ws.row_dimensions[13].height = 18
    ws.row_dimensions[14].height = 18

    ws["D1"].value = "Bill To:"
    ws["D1"].font = bold
    ws.merge_cells("D1:F1")

    ws["H1"].value = "Ship To:"
    ws["H1"].font = bold
    ws.merge_cells("H1:J1")

    bill_text = "\n".join([x for x in [
        (bill_to.get("name") or "").strip(),
        (bill_to.get("address") or "").strip(),
        (bill_to.get("address2") or "").strip()
    ] if x])

    ship_text = "\n".join([x for x in [
        (ship_to.get("name") or "").strip(),
        (ship_to.get("address") or "").strip(),
        (ship_to.get("address2") or "").strip()
    ] if x])

    ws["D2"].value = bill_text
    ws.merge_cells("D2:F3")
    ws["D2"].alignment = left

    ws["H2"].value = ship_text
    ws.merge_cells("H2:J3")
    ws["H2"].alignment = left

    ws["D5"].value = "Phone:"
    ws["D5"].font = bold
    ws.merge_cells("E5:F5")
    ws["E5"].value = phone
    ws["E5"].alignment = left_mid

    ws["H5"].value = "Fax:"
    ws["H5"].font = bold
    ws.merge_cells("I5:J5")
    ws["I5"].value = fax
    ws["I5"].alignment = left_mid

    ws["D7"].value = "Attn :"
    ws["D7"].font = bold
    ws.merge_cells("E7:F7")
    ws["E7"].value = attn
    ws["E7"].alignment = left_mid

    ws["I6"].value = "Invoice"
    ws["I6"].font = bold
    ws["I6"].alignment = right_mid
    ws["J6"].value = invoice_no
    ws["J6"].alignment = left_mid

    ws["I7"].value = "Date"
    ws["I7"].font = bold
    ws["I7"].alignment = right_mid
    ws["J7"].value = inv_date
    ws["J7"].alignment = left_mid

    ws["I8"].value = "No. Surat Jalan"
    ws["I8"].font = bold
    ws["I8"].alignment = right_mid
    ws["J8"].value = no_surat_jalan
    ws["J8"].alignment = left_mid

    ws.merge_cells("D10:E10")
    ws["D10"].value = "Ref No."
    ws["D10"].font = bold
    ws["D10"].alignment = center

    ws.merge_cells("F10:G10")
    ws["F10"].value = "Sales Person"
    ws["F10"].font = bold
    ws["F10"].alignment = center

    ws["H10"].value = "Ship Via"
    ws["H10"].font = bold
    ws["H10"].alignment = center

    ws["I10"].value = "Ship Date"
    ws["I10"].font = bold
    ws["I10"].alignment = center

    ws["J10"].value = "Terms"
    ws["J10"].font = bold
    ws["J10"].alignment = center

    ws.merge_cells("D11:E13")
    ws["D11"].value = ref_no
    ws["D11"].alignment = center

    ws.merge_cells("F11:G13")
    ws["F11"].value = sales_person
    ws["F11"].alignment = center

    ws.merge_cells("H11:H13")
    ws["H11"].value = ship_via
    ws["H11"].alignment = center

    ws.merge_cells("I11:I13")
    ws["I11"].value = ship_date
    ws["I11"].alignment = center

    ws.merge_cells("J11:J13")
    ws["J11"].value = terms
    ws["J11"].alignment = center

    apply_outer_and_vertical_only(ws, 10, 4, 13, 10, vertical_separators_cols=[6, 8, 9, 10])

    ws["D14"].value = "Qty"
    ws["E14"].value = ""
    ws["F14"].value = "Date"
    ws.merge_cells("G14:H14")
    ws["G14"].value = "Description"
    ws["I14"].value = "Price"
    ws["J14"].value = "Amount (IDR)"

    for c in ["D", "E", "F", "G", "I", "J"]:
        ws[f"{c}14"].font = bold
        ws[f"{c}14"].alignment = center
    ws["H14"].alignment = center

    items = inv.get("items") or []
    start_row = 15
    max_rows = max(10, len(items))
    subtotal = 0

    for idx in range(max_rows):
        r = start_row + idx
        ws.merge_cells(f"G{r}:H{r}")

        ws[f"D{r}"].alignment = center
        ws[f"E{r}"].alignment = center
        ws[f"F{r}"].alignment = center
        ws[f"G{r}"].alignment = left
        ws[f"I{r}"].alignment = right
        ws[f"J{r}"].alignment = right

        if idx < len(items):
            it = items[idx]
            qty = float(it.get("qty") or 0)
            unit = (it.get("unit") or "Kg").strip()
            dt = it.get("date") or inv_date
            desc = (it.get("description") or "").strip()
            price = int(it.get("price") or 0)
            amount = int(round(qty * price))
            subtotal += amount

            ws[f"D{r}"].value = qty if qty % 1 != 0 else int(qty)
            ws[f"E{r}"].value = unit
            ws[f"F{r}"].value = dt
            ws[f"G{r}"].value = desc
            ws[f"I{r}"].value = price
            ws[f"J{r}"].value = amount
            money(ws[f"I{r}"])
            money(ws[f"J{r}"])

    last_table_row = start_row + max_rows - 1
    apply_outer_and_vertical_only(ws, 14, 4, last_table_row, 10, vertical_separators_cols=[5, 6, 7, 9, 10])

    freight = int(inv.get("freight") or 0)
    ppn_rate = float(inv.get("ppn_rate") or 0.11)
    deposit = int(inv.get("deposit") or 0)

    total_before_ppn = subtotal + freight
    ppn = int(round(total_before_ppn * ppn_rate))
    balance = total_before_ppn + ppn - deposit

    base_row = last_table_row + 2

    ws.merge_cells(f"D{base_row}:H{base_row}")
    ws[f"D{base_row}"].value = "Please Transfer Full Amount to:"
    ws[f"D{base_row}"].font = bold_ul
    ws[f"D{base_row}"].alignment = left_mid

    left_lines = [
        f"Beneficiary : {payment.get('beneficiary','')}",
        f"Bank Name   : {payment.get('bank_name','')}",
        f"Branch      : {payment.get('branch','')}",
        f"IDR Acct    : {payment.get('idr_acct','')}",
    ]
    for i, line in enumerate(left_lines, start=1):
        ws.merge_cells(f"D{base_row+i}:H{base_row+i}")
        ws[f"D{base_row+i}"].value = line
        ws[f"D{base_row+i}"].alignment = left_mid

    labels = [
        ("Total", subtotal),
        ("Freight", freight),
        ("Total", total_before_ppn),
        (f"PPN {int(ppn_rate*100)}%", ppn),
        ("Less: Deposit", deposit),
        ("Balance Due", balance),
    ]
    totals_top = base_row

    thin_side = Side(style="thin", color="000000")
    no_border = Border()

    def border_box_thin():
        return Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for i, (lab, val) in enumerate(labels):
        rr = totals_top + i

        ws[f"I{rr}"].value = lab
        ws[f"I{rr}"].alignment = right
        ws[f"I{rr}"].font = bold
        ws[f"I{rr}"].border = no_border

        ws[f"J{rr}"].value = val
        ws[f"J{rr}"].alignment = right
        ws[f"J{rr}"].font = bold if lab in ("Balance Due",) else Font(bold=False)
        ws[f"J{rr}"].border = border_box_thin()
        money(ws[f"J{rr}"])

    totals_bottom = totals_top + len(labels) - 1

    box_top = totals_bottom + 2
    box_bottom = box_top + 6

    ws.merge_cells(f"G{box_top}:J{box_top}")
    ws[f"G{box_top}"].value = "PT. Sarana Trans Bersama Jaya"
    ws[f"G{box_top}"].alignment = center
    ws[f"G{box_top}"].font = bold

    set_outer_border_only(ws, box_top, 7, box_bottom, 10, style="medium")

    footer_row = box_bottom + 1
    ws.merge_cells(f"G{footer_row}:J{footer_row}")
    ws[f"G{footer_row}"].value = "Please kindly fax to our attention upon receipt"
    ws[f"G{footer_row}"].alignment = center

    try:
        folder = str(FILES_DIR)
    except Exception:
        folder = "static/files"
    os.makedirs(folder, exist_ok=True)

    out_path = os.path.join(folder, f"{fname_base}.xlsx")
    wb.save(out_path)
    return f"{fname_base}.xlsx"


# ==========================================
# PDF: biarkan seperti kode Anda (tidak saya ubah)
# ==========================================
def create_invoice_pdf(inv: dict, fname_base: str) -> str:
    try:
        folder = str(FILES_DIR)
    except Exception:
        folder = "static/files"
    os.makedirs(folder, exist_ok=True)

    out_path = os.path.join(folder, f"{fname_base}.pdf")
    c = canvas.Canvas(out_path, pagesize=A4)
    width, height = A4

    def txt(x, y, s, size=9, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(x, y, s or "")

    def rtxt(x, y, s, size=9, bold=False):
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawRightString(x, y, s or "")

    def rect(x, y, w, h, lw=1):
        c.setLineWidth(lw)
        c.rect(x, y, w, h)

    def vline(x, y1, y2, lw=0.6):
        c.setLineWidth(lw)
        c.line(x, y1, x, y2)

    def fmt_id(n: int) -> str:
        try:
            return f"{int(n):,}".replace(",", ".")
        except:
            return str(n)

    invoice_no = inv.get("invoice_no") or ""
    inv_date = inv.get("invoice_date") or ""
    bill_to = inv.get("bill_to") or {}
    ship_to = inv.get("ship_to") or {}
    phone = inv.get("phone") or ""
    fax = inv.get("fax") or ""
    attn = inv.get("attn") or "Accounting / Finance"
    sales_person = inv.get("sales_person") or "Syaeful Bakri"
    ref_no = inv.get("ref_no") or ""
    ship_via = inv.get("ship_via") or ""
    ship_date = inv.get("ship_date") or ""
    terms = inv.get("terms") or ""
    no_surat_jalan = inv.get("no_surat_jalan") or ""
    items = inv.get("items") or []
    freight = int(inv.get("freight") or 0)
    ppn_rate = float(inv.get("ppn_rate") or 0.11)
    deposit = int(inv.get("deposit") or 0)
    payment = inv.get("payment") or {}

    left_margin = 40
    table_x = left_margin
    table_w = width - 80

    w_qty = 45
    w_unit = 35
    w_date = 70
    w_desc = 220
    w_price = 70
    w_amt = table_w - (w_qty + w_unit + w_date + w_desc + w_price)

    x_qty = table_x
    x_unit = x_qty + w_qty
    x_date = x_unit + w_unit
    x_desc = x_date + w_date
    x_price = x_desc + w_desc
    x_amt = x_price + w_price
    x_end = table_x + table_w

    y = height - 50

    txt(table_x, y, "Bill To:", 10, True)
    txt(table_x + table_w * 0.55, y, "Ship To:", 10, True)
    y -= 14

    bt_lines = [bill_to.get("name", ""), bill_to.get("address", ""), bill_to.get("address2", "")]
    st_lines = [ship_to.get("name", ""), ship_to.get("address", ""), ship_to.get("address2", "")]
    bt_lines = [s for s in bt_lines if (s or "").strip()]
    st_lines = [s for s in st_lines if (s or "").strip()]

    yy = y
    for line in bt_lines[:3]:
        txt(table_x, yy, str(line), 9, False)
        yy -= 12

    yy2 = y
    for line in st_lines[:3]:
        txt(table_x + table_w * 0.55, yy2, str(line), 9, False)
        yy2 -= 12

    rtxt(x_end, height - 62, invoice_no, 9, False)
    txt(x_end - 120, height - 62, "Invoice", 9, True)
    rtxt(x_end, height - 76, inv_date, 9, False)
    txt(x_end - 120, height - 76, "Date", 9, True)
    rtxt(x_end, height - 90, no_surat_jalan, 9, False)
    txt(x_end - 120, height - 90, "No. Surat Jalan", 9, True)

    y = min(yy, yy2) - 8
    txt(table_x, y, "Phone:", 9, True)
    txt(table_x + 50, y, phone, 9, False)
    txt(table_x + table_w * 0.55, y, "Fax:", 9, True)
    txt(table_x + table_w * 0.55 + 35, y, fax, 9, False)
    y -= 14
    txt(table_x, y, "Attn :", 9, True)
    txt(table_x + 45, y, attn, 9, False)

    y -= 28
    ref_box_top = y
    ref_box_h = 40
    rect(table_x, ref_box_top - ref_box_h, table_w, ref_box_h, lw=1)

    vline(table_x + table_w * 0.25, ref_box_top - ref_box_h, ref_box_top, lw=0.6)
    vline(table_x + table_w * 0.55, ref_box_top - ref_box_h, ref_box_top, lw=0.6)
    vline(table_x + table_w * 0.78, ref_box_top - ref_box_h, ref_box_top, lw=0.6)

    txt(table_x + 10, ref_box_top - 14, "Ref No.", 9, True)
    txt(table_x + table_w * 0.25 + 10, ref_box_top - 14, "Sales Person", 9, True)
    txt(table_x + table_w * 0.55 + 10, ref_box_top - 14, "Ship Via", 9, True)
    txt(table_x + table_w * 0.78 + 10, ref_box_top - 14, "Ship Date", 9, True)

    txt(table_x + 10, ref_box_top - 30, ref_no, 9, False)
    txt(table_x + table_w * 0.25 + 10, ref_box_top - 30, sales_person, 9, False)
    txt(table_x + table_w * 0.55 + 10, ref_box_top - 30, ship_via, 9, False)
    txt(table_x + table_w * 0.78 + 10, ref_box_top - 30, ship_date, 9, False)

    txt(x_amt - 5, ref_box_top - ref_box_h - 14, "Terms", 9, True)
    rtxt(x_end, ref_box_top - ref_box_h - 14, terms, 9, False)

    y = ref_box_top - ref_box_h - 28
    table_top = y
    table_h = 220
    rect(table_x, table_top - table_h, table_w, table_h, lw=1)

    vline(x_unit, table_top - table_h, table_top, lw=0.6)
    vline(x_date, table_top - table_h, table_top, lw=0.6)
    vline(x_desc, table_top - table_h, table_top, lw=0.6)
    vline(x_price, table_top - table_h, table_top, lw=0.6)
    vline(x_amt, table_top - table_h, table_top, lw=0.6)

    header_y = table_top - 16
    txt(x_qty + 4, header_y, "Qty", 9, True)
    txt(x_date + 4, header_y, "Date", 9, True)
    txt(x_desc + 4, header_y, "Description", 9, True)
    txt(x_price + 4, header_y, "Price", 9, True)
    txt(x_amt + 4, header_y, "Amount (IDR)", 9, True)

    row_y = header_y - 18
    subtotal = 0
    max_rows = 10
    for idx in range(max_rows):
        if idx < len(items):
            it = items[idx]
            qty = it.get("qty") or 0
            unit = it.get("unit") or "Kg"
            dt = it.get("date") or inv_date
            desc = it.get("description") or ""
            price = int(it.get("price") or 0)
            amount = int(round(float(qty) * price))
            subtotal += amount

            txt(x_qty + 4, row_y, str(qty), 9, False)
            txt(x_unit + 4, row_y, str(unit), 9, False)
            txt(x_date + 4, row_y, str(dt), 9, False)
            txt(x_desc + 4, row_y, str(desc)[:45], 9, False)
            rtxt(x_price + w_price - 4, row_y, fmt_id(price), 9, False)
            rtxt(x_end - 4, row_y, fmt_id(amount), 9, False)
        row_y -= 16

    total_before_ppn = subtotal + freight
    ppn = int(round(total_before_ppn * ppn_rate))
    balance = total_before_ppn + ppn - deposit

    base_y = table_top - table_h - 20

    txt(table_x, base_y, "Please Transfer Full Amount to:", 9, True)
    txt(table_x, base_y - 14, f"Beneficiary : {payment.get('beneficiary','')}", 9, False)
    txt(table_x, base_y - 28, f"Bank Name   : {payment.get('bank_name','')}", 9, False)
    txt(table_x, base_y - 42, f"Branch      : {payment.get('branch','')}", 9, False)
    txt(table_x, base_y - 56, f"IDR Acct    : {payment.get('idr_acct','')}", 9, False)

    box_w = w_price + w_amt
    box_x = x_price
    box_y_top = base_y + 8
    line_h = 14
    labels = [
        ("Total", subtotal),
        ("Freight", freight),
        ("Total", total_before_ppn),
        (f"PPN {int(ppn_rate*100)}%", ppn),
        ("Less: Deposit", deposit),
        ("Balance Due", balance),
    ]
    box_h = line_h * len(labels) + 6
    rect(box_x, box_y_top - box_h, box_w, box_h, lw=1)

    yy = box_y_top - 16
    for (lab, val) in labels:
        txt(box_x + 6, yy, lab, 9, True if lab in ("Total", "Balance Due") else False)
        rtxt(box_x + box_w - 6, yy, fmt_id(val), 9, True if lab in ("Balance Due",) else False)
        yy -= line_h

    sig_top = box_y_top - box_h - 30
    sig_w = box_w
    sig_h = 80
    rect(box_x, sig_top - sig_h, sig_w, sig_h, lw=1)
    txt(box_x + 10, sig_top - 14, "PT. Sarana Trans Bersama Jaya", 9, True)

    txt(box_x + 10, sig_top - sig_h - 14, "Please kindly fax to our attention upon receipt", 9, False)

    c.showPage()
    c.save()
    return f"{fname_base}.pdf"


def handle_invoice_flow(data: dict, text: str, lower: str, sid: str, state: dict, conversations: dict, history_id_in):
    text = normalize_voice_strip(text)
    lower = (text or "").strip().lower()

    if (("invoice" in lower) or ("faktur" in lower)) and (state.get("step") == "idle"):
        inv_no = get_next_invoice_no()
        state["step"] = "inv_billto_name"
        state["data"] = {
            "invoice_no": inv_no,
            "invoice_date": datetime.now().strftime("%d-%b-%y"),
            "bill_to": {"name": "", "address": "", "address2": ""},
            "ship_to": {"name": "", "address": "", "address2": ""},
            "phone": "",
            "fax": "",
            "attn": "Accounting / Finance",
            "sales_person": "Syaeful Bakri",
            "ref_no": "",
            "ship_via": "",
            "ship_date": "",
            "terms": "",
            "no_surat_jalan": "",
            "items": [],
            "current_item": {},
            "freight": 0,
            "ppn_rate": 0.11,
            "deposit": 0,
            "payment": {
                "beneficiary": "PT. Sarana Trans Bersama Jaya",
                "bank_name": "BCA",
                "branch": "Cibadak - Sukabumi",
                "idr_acct": "35212 26666",
            }
        }
        conversations[sid] = state

        out_text = (
            "Baik, saya akan membantu membuat <b>Invoice</b>.<br><br>"
            f"Nomor invoice: <b>{inv_no}</b><br>"
            f"Tanggal: <b>{state['data']['invoice_date']}</b><br><br>"
            "Pertanyaan 1: <b>Nama perusahaan untuk Bill To?</b>"
        )

        history_id_created = None
        if not history_id_in:
            history_id_created = db_insert_history(
                title="Chat Baru",
                task_type=data.get("taskType") or "invoice",
                data={},
                files=[],
                messages=[
                    {"id": __import__("uuid").uuid4().hex[:12], "sender": "user", "text": text, "files": [], "timestamp": datetime.now().isoformat()},
                    {"id": __import__("uuid").uuid4().hex[:12], "sender": "assistant", "text": re.sub(r"<br\s*/?>", "\n", out_text), "files": [], "timestamp": datetime.now().isoformat()},
                ],
                state=state
            )
        else:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)

        return {"text": out_text, "history_id": history_id_created or history_id_in}

    if state.get("step") == "inv_billto_name":
        state["data"]["bill_to"]["name"] = text.strip()
        alamat = resolve_company_address(text)
        state["data"]["bill_to"]["address"] = alamat
        state["step"] = "inv_shipto_same"
        conversations[sid] = state

        out_text = (
            f"Bill To: <b>{state['data']['bill_to']['name']}</b><br>"
            f"Alamat: <b>{alamat}</b><br><br>"
            "Pertanyaan 2: <b>Apakah Ship To sama dengan Bill To?</b> (ya/tidak)"
        )

        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_shipto_same":
        if ("ya" in lower) or ("iya" in lower):
            state["data"]["ship_to"] = dict(state["data"]["bill_to"])
            state["step"] = "inv_phone"
            conversations[sid] = state
            out_text = "Pertanyaan 3: <b>Nomor telepon?</b> (boleh kosong; sebut <b>strip</b> jika tidak ada)"
        elif ("tidak" in lower) or ("gak" in lower) or ("nggak" in lower):
            state["step"] = "inv_shipto_name"
            conversations[sid] = state
            out_text = "Pertanyaan 2A: <b>Nama perusahaan untuk Ship To?</b>"
        else:
            out_text = (
                "Mohon jawab dengan <b>ya</b> atau <b>tidak</b>.<br><br>"
                "Pertanyaan 2: <b>Apakah Ship To sama dengan Bill To?</b>"
            )
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_shipto_name":
        state["data"]["ship_to"]["name"] = text.strip()
        alamat = resolve_company_address(text)
        state["data"]["ship_to"]["address"] = alamat
        state["step"] = "inv_phone"
        conversations[sid] = state
        out_text = "Pertanyaan 3: <b>Nomor telepon?</b> (boleh kosong; sebut <b>strip</b> jika tidak ada)"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_phone":
        state["data"]["phone"] = "" if text.strip() in ("-", "") else text.strip()
        state["step"] = "inv_fax"
        conversations[sid] = state
        out_text = "Pertanyaan 4: <b>Fax?</b> (boleh kosong; sebut <b>strip</b> jika tidak ada)"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_fax":
        state["data"]["fax"] = "" if text.strip() in ("-", "") else text.strip()
        state["step"] = "inv_attn"
        conversations[sid] = state
        out_text = "Pertanyaan 5: <b>Attn?</b> (default: Accounting / Finance; sebut <b>strip</b> untuk default)"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_attn":
        if text.strip() not in ("-", ""):
            state["data"]["attn"] = text.strip()

        state["step"] = "inv_item_qty"
        state["data"]["current_item"] = {}
        conversations[sid] = state
        out_text = (
            "Item 1<br>"
            "Pertanyaan 6: <b>Qty?</b> (contoh: 749 atau 2,5 atau 'dua koma lima')"
        )
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_item_qty":
        qty = parse_qty_id(text)
        state["data"]["current_item"]["qty"] = qty
        state["data"]["current_item"]["unit"] = "Kg"
        state["data"]["current_item"]["date"] = state["data"]["invoice_date"]

        state["step"] = "inv_item_desc"
        conversations[sid] = state
        out_text = (
            "Pertanyaan 6B: <b>Jenis limbah atau kode limbah?</b><br>"
            "<i>Contoh: A102d atau aki baterai bekas. Atau sebut <b>NON B3</b>.</i>"
        )
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_item_desc":
        if is_non_b3_input(text):
            state["data"]["current_item"]["description"] = ""
            state["step"] = "inv_item_desc_manual"
            conversations[sid] = state
            out_text = "Pertanyaan 6C: <b>Deskripsi (manual) apa?</b>"
            if history_id_in:
                db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
                db_update_state(int(history_id_in), state)
            return {"text": out_text, "history_id": history_id_in}

        kode, data_limbah = find_limbah_by_kode(text)
        if not (kode and data_limbah):
            kode, data_limbah = find_limbah_by_jenis(text)

        if kode and data_limbah:
            state["data"]["current_item"]["description"] = data_limbah["jenis"]
            state["step"] = "inv_item_price"
            conversations[sid] = state
            out_text = (
                f"Deskripsi: <b>{data_limbah['jenis']}</b><br><br>"
                "Pertanyaan 6D: <b>Harga (Rp)?</b>"
            )
            if history_id_in:
                db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
                db_update_state(int(history_id_in), state)
            return {"text": out_text, "history_id": history_id_in}

        out_text = (
            f"Maaf, limbah <b>{text}</b> tidak ditemukan.<br><br>"
            "Silakan sebutkan kode/jenis lain atau sebut <b>NON B3</b> untuk input manual."
        )
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_item_desc_manual":
        state["data"]["current_item"]["description"] = text.strip()
        state["step"] = "inv_item_price"
        conversations[sid] = state
        out_text = "Pertanyaan 6D: <b>Harga (Rp)?</b>"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_item_price":
        price = parse_amount_id(text)
        state["data"]["current_item"]["price"] = price
        state["data"]["items"].append(state["data"]["current_item"])
        state["data"]["current_item"] = {}
        state["step"] = "inv_add_more_item"
        conversations[sid] = state
        out_text = "Pertanyaan: <b>Tambah item lagi?</b> (ya/tidak)"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_add_more_item":
        if ("ya" in lower) or ("iya" in lower):
            num = len(state["data"]["items"])
            state["step"] = "inv_item_qty"
            state["data"]["current_item"] = {}
            conversations[sid] = state
            out_text = f"Item {num+1}<br>Pertanyaan 6: <b>Qty?</b>"
            if history_id_in:
                db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
                db_update_state(int(history_id_in), state)
            return {"text": out_text, "history_id": history_id_in}

        if ("tidak" in lower) or ("gak" in lower) or ("nggak" in lower) or ("skip" in lower) or ("lewat" in lower):
            state["step"] = "inv_freight"
            conversations[sid] = state
            out_text = "Pertanyaan 7: <b>Biaya transportasi/Freight (Rp)?</b> (isi 0 jika tidak ada)"
            if history_id_in:
                db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
                db_update_state(int(history_id_in), state)
            return {"text": out_text, "history_id": history_id_in}

        out_text = "Mohon jawab <b>ya</b> atau <b>tidak</b>."
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_freight":
        state["data"]["freight"] = parse_amount_id(text)
        state["step"] = "inv_deposit"
        conversations[sid] = state
        out_text = "Pertanyaan 8: <b>Deposit (Rp)?</b> (isi 0 jika tidak ada)"
        if history_id_in:
            db_append_message(int(history_id_in), "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=[])
            db_update_state(int(history_id_in), state)
        return {"text": out_text, "history_id": history_id_in}

    if state.get("step") == "inv_deposit":
        state["data"]["deposit"] = parse_amount_id(text)

        nama_pt_raw = (state["data"].get("bill_to") or {}).get("name", "").strip()
        safe_pt = re.sub(r"[^A-Za-z0-9 \-]+", "", nama_pt_raw).strip()
        safe_pt = re.sub(r"\s+", " ", safe_pt).strip()
        base_fname = f"Invoice - {safe_pt}" if safe_pt else "Invoice"
        fname_base = make_unique_filename_base(base_fname)

        xlsx = create_invoice_xlsx(state["data"], fname_base)
        pdf_preview = create_invoice_pdf(state["data"], fname_base)

        files = [
            {"type": "xlsx", "filename": xlsx, "url": f"/download/{xlsx}"},
            {"type": "pdf", "filename": pdf_preview, "url": f"/download/{pdf_preview}"},
        ]

        conversations[sid] = {"step": "idle", "data": {}}

        history_title = f"Invoice {nama_pt_raw}" if nama_pt_raw else "Invoice"
        history_task_type = "invoice"

        if history_id_in:
            from utils import db_connect
            conn = db_connect()
            cur = conn.cursor()
            cur.execute("""
                UPDATE chat_history
                SET title = ?, task_type = ?, data_json = ?, files_json = ?
                WHERE id = ?
            """, (
                history_title,
                history_task_type,
                json.dumps(state["data"], ensure_ascii=False),
                json.dumps(files, ensure_ascii=False),
                int(history_id_in),
            ))
            conn.commit()
            conn.close()
            history_id = int(history_id_in)
        else:
            history_id = db_insert_history(
                title=history_title,
                task_type=history_task_type,
                data=state["data"],
                files=files,
                messages=[],
                state={}
            )

        out_text = (
            "<b>Invoice berhasil dibuat.</b><br><br>"
            f"Nomor invoice: <b>{state['data'].get('invoice_no')}</b><br>"
            f"Bill To: <b>{(state['data'].get('bill_to') or {}).get('name','')}</b><br>"
            f"Jumlah item: <b>{len(state['data'].get('items') or [])}</b><br><br>"
            "Dokumen tersedia dalam format PDF (preview) dan Excel (.xlsx)."
        )

        db_append_message(history_id, "assistant", re.sub(r"<br\s*/?>", "\n", out_text), files=files)
        return {"text": out_text, "files": files, "history_id": history_id}

    return None
